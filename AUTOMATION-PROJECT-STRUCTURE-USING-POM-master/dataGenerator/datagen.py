import openpyxl

def datagenerator():
    # using list for taking data
    #li = [['uname1','email1','pass1'],['uname2','email2','pass2'],['uname3','email3','pass3']]
    #return li
    #usong xl sheet
    wk = openpyxl.load_workbook("D:/software testing/ts_001.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    li =[]
    li1 =[]
    for i in range(1,r+1):
        li1=[]
        un =sh.cell(i,1)
        ue =sh.cell(i,2)
        up =sh.cell(i,3)
        li1.insert(0, un.value)
        li1.insert(1, ue.value)
        li1.insert(2, up.value)
        li.insert(i-1,li1)
    print(li)
    return li
